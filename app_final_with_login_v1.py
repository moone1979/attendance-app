import streamlit as st
import pandas as pd
import os
import time
import re
import io
import zipfile
import zoneinfo
import streamlit.components.v1 as components
import math
from datetime import datetime, date, timedelta

# 日本時間のタイムゾーン設定
JST = zoneinfo.ZoneInfo("Asia/Tokyo")
def today_jst():
    return datetime.now(JST).date()

st.set_page_config(page_title="出退勤アプリ（ログイン式）", layout="wide")

st.markdown("""
<style>
.material-icons,
.material-icons-outlined,
.material-icons-round,
.material-icons-sharp,
.material-icons-two-tone {
  font-family: 'Material Icons' !important;
  font-weight: normal !important;
  font-style: normal !important;
  font-size: inherit;
  line-height: 1;
  -webkit-font-feature-settings: 'liga';
  -webkit-font-smoothing: antialiased;
}
</style>
""", unsafe_allow_html=True)

# ==============================
# パス & 列定義
# ==============================
DATA_DIR = os.getenv("DATA_DIR", ".")
CSV_PATH      = os.path.join(DATA_DIR, "attendance_log.csv")
LOGIN_CSV     = os.path.join(DATA_DIR, "社員ログイン情報.csv")
HOLIDAY_CSV   = os.path.join(DATA_DIR, "holiday_requests.csv")
AUDIT_LOG_CSV = os.path.join(DATA_DIR, "holiday_audit_log.csv")
OVERTIME_CSV = os.path.join(DATA_DIR, "overtime_requests.csv")

LOGIN_COLUMNS   = ["社員ID", "氏名", "部署", "パスワード"]
ATT_COLUMNS     = ["社員ID", "氏名", "日付", "出勤時刻", "退勤時刻", "緯度", "経度"]
HOLIDAY_COLUMNS = ["社員ID", "氏名", "申請日", "休暇日", "休暇種類", "備考", "ステータス", "承認者", "承認日時", "却下理由"]
AUDIT_COLUMNS   = ["timestamp","承認者","社員ID","氏名","休暇日","申請日","旧ステータス","新ステータス","却下理由"]
OVERTIME_COLUMNS = ["社員ID","氏名","対象日","申請日時","申請残業H","申請理由","ステータス","承認者","承認日時","却下理由"]

os.makedirs(DATA_DIR, exist_ok=True)

# ==============================
# CSV初期化
# ==============================
if not os.path.exists(CSV_PATH):
    pd.DataFrame(columns=ATT_COLUMNS).to_csv(CSV_PATH, index=False, encoding="utf-8-sig")

if not os.path.exists(HOLIDAY_CSV):
    pd.DataFrame(columns=HOLIDAY_COLUMNS).to_csv(HOLIDAY_CSV, index=False, encoding="utf-8-sig")

if not os.path.exists(OVERTIME_CSV):
    pd.DataFrame(columns=OVERTIME_COLUMNS).to_csv(OVERTIME_CSV, index=False, encoding="utf-8-sig")

# ==============================
# UTF-8 修復
# ==============================
def _read_csv_flexible(path: str) -> pd.DataFrame:
    for enc in ("utf-8-sig", "utf-8", "cp932"):
        try:
            return pd.read_csv(path, dtype=str, encoding=enc).fillna("")
        except UnicodeDecodeError:
            continue
    return pd.read_csv(path, dtype=str, encoding="cp932", encoding_errors="replace").fillna("")

def safe_write_csv(df: pd.DataFrame, path: str, columns: list[str], retries=5, wait=0.8):
    for _ in range(retries):
        try:
            _write_atomic_csv(df, path, columns)  # *.tmp → os.replace で安全置換
            return True
        except PermissionError:
            time.sleep(wait)
    st.error("CSVを書き込めません。Excel/プレビュー/同期を閉じてから再実行してください。")
    return False

# ==============================
# CSVインジェクション対策（Excelでの式実行防止）
# ==============================
def sanitize_for_csv(value: str) -> str:
    """
    セルの先頭が Excel 式 (=, +, -, @) と解釈されるのを防止する
    """
    if not isinstance(value, str):
        return value
    if value.startswith(("=", "+", "-", "@")):
        return "'" + value  # シングルクォートで無害化
    return value

# ==============================
# 残業申請 CSV 操作
# ==============================
def read_overtime_csv() -> pd.DataFrame:
    if not os.path.exists(OVERTIME_CSV):
        df = pd.DataFrame(columns=OVERTIME_COLUMNS)
        df.to_csv(OVERTIME_CSV, index=False, encoding="utf-8-sig")
        return df.copy()
    for enc in ("utf-8-sig", "utf-8", "cp932"):
        try:
            df = pd.read_csv(OVERTIME_CSV, dtype=str, encoding=enc).fillna("")
            break
        except UnicodeDecodeError:
            continue
    else:
        df = pd.read_csv(OVERTIME_CSV, dtype=str, encoding="cp932", encoding_errors="replace").fillna("")
    for col in OVERTIME_COLUMNS:
        if col not in df.columns:
            df[col] = ""
    return df[OVERTIME_COLUMNS].copy()

def write_overtime_csv(df: pd.DataFrame):
    df = df.applymap(sanitize_for_csv)
    for col in OVERTIME_COLUMNS:
        if col not in df.columns:
            df[col] = ""
    safe_write_csv(df[OVERTIME_COLUMNS], OVERTIME_CSV, OVERTIME_COLUMNS)

# ==============================
# 休日申請 CSV 操作
# ==============================
def read_holiday_csv() -> pd.DataFrame:
    if not os.path.exists(HOLIDAY_CSV):
        df = pd.DataFrame(columns=HOLIDAY_COLUMNS)
        df.to_csv(HOLIDAY_CSV, index=False, encoding="utf-8-sig")
        return df.copy()
    for enc in ("utf-8-sig", "utf-8", "cp932"):
        try:
            df = pd.read_csv(HOLIDAY_CSV, dtype=str, encoding=enc).fillna("")
            break
        except UnicodeDecodeError:
            continue
    else:
        df = pd.read_csv(HOLIDAY_CSV, dtype=str, encoding="cp932", encoding_errors="replace").fillna("")
    for col in HOLIDAY_COLUMNS:
        if col not in df.columns:
            df[col] = ""
    return df[HOLIDAY_COLUMNS].copy()

def write_holiday_csv(df: pd.DataFrame):
    # --- CSVインジェクション対策を適用 ---
    df = df.applymap(sanitize_for_csv)

    for col in HOLIDAY_COLUMNS:
        if col not in df.columns:
            df[col] = ""
    safe_write_csv(df[HOLIDAY_COLUMNS], HOLIDAY_CSV, HOLIDAY_COLUMNS)

# ==============================
# 監査ログユーティリティ
# ==============================
def append_audit_log(rows: list[dict]):
    if not rows: return
    file_exists = os.path.exists(AUDIT_LOG_CSV)
    pd.DataFrame(rows, columns=AUDIT_COLUMNS).to_csv(
        AUDIT_LOG_CSV, index=False, encoding="utf-8-sig", mode="a", header=not file_exists
    )

# 勤怠入力で「申請済」を自動取消（監査ログは system）
def auto_cancel_holiday_by_attendance(user_id: str, user_name: str, work_date_str: str) -> int:
    hd = read_holiday_csv()
    if hd.empty: return 0
    mask = (
        (hd["社員ID"].astype(str) == str(user_id)) &
        (hd["休暇日"] == work_date_str) &
        (hd["ステータス"] == "申請済")
    )
    cnt = int(mask.sum())
    if cnt == 0: return 0
    ts = datetime.now(JST).strftime("%Y-%m-%d %H:%M:%S")
    rows = []
    for _, r in hd[mask].iterrows():
        rows.append({
            "timestamp": ts, "承認者": "system",
            "社員ID": user_id, "氏名": user_name,
            "休暇日": work_date_str, "申請日": r.get("申請日",""),
            "旧ステータス": "申請済", "新ステータス": "自動取消(勤怠入力)", "却下理由": ""
        })
    hd2 = hd[~mask].copy()
    write_holiday_csv(hd2)
    append_audit_log(rows)
    return cnt

# ==============================
# バックアップ/復元ヘルパー
# ==============================
def _read_existing_or_empty(path: str, columns: list[str]) -> pd.DataFrame:
    if os.path.exists(path):
        for enc in ("utf-8-sig", "utf-8", "cp932"):
            try:
                return pd.read_csv(path, dtype=str, encoding=enc).fillna("")
            except UnicodeDecodeError:
                continue
        return pd.read_csv(path, dtype=str, encoding="cp932", encoding_errors="replace").fillna("")
    else:
        return pd.DataFrame(columns=columns)

def _write_atomic_csv(df: pd.DataFrame, path: str, columns: list[str]):
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    for c in columns:
        if c not in df.columns:
            df[c] = ""
    tmp = path + ".tmp"
    df[columns].to_csv(tmp, index=False, encoding="utf-8-sig")
    os.replace(tmp, path)

def _read_csv_bytes(data: bytes) -> pd.DataFrame:
    for enc in ("utf-8-sig", "utf-8", "cp932"):
        try:
            return pd.read_csv(io.BytesIO(data), dtype=str, encoding=enc).fillna("")
        except UnicodeDecodeError:
            continue
    return pd.read_csv(io.BytesIO(data), dtype=str, encoding="cp932", encoding_errors="replace").fillna("")

# 期待するファイル名と対応付け（エクスポート/インポートで共通）
BACKUP_TABLES = [
    (CSV_PATH,      ATT_COLUMNS,     "attendance_log.csv"),
    (HOLIDAY_CSV,   HOLIDAY_COLUMNS, "holiday_requests.csv"),
    (AUDIT_LOG_CSV, AUDIT_COLUMNS,   "holiday_audit_log.csv"),
    (LOGIN_CSV,     LOGIN_COLUMNS,   "社員ログイン情報.csv"),
]

# ==============================
# 社員ログイン情報 救済
# ==============================
def read_login_csv(path: str) -> pd.DataFrame:
    if not os.path.exists(path):
        return pd.DataFrame(columns=LOGIN_COLUMNS)
    for enc in ("utf-8-sig", "utf-8", "cp932"):
        try:
            df = pd.read_csv(path, dtype=str, encoding=enc).fillna("")
            break
        except UnicodeDecodeError:
            continue
    else:
        df = pd.read_csv(path, dtype=str, encoding="cp932", encoding_errors="replace").fillna("")
    for col in LOGIN_COLUMNS:
        if col not in df.columns:
            df[col] = ""
    return df[LOGIN_COLUMNS].astype({"社員ID":str,"氏名":str,"部署":str,"パスワード":str}).copy()

df_login = read_login_csv(LOGIN_CSV)

# === クエリからの自動ログイン（一般社員のみ） ===
qs = st.query_params
uid_q = qs.get("uid")
if uid_q and not st.session_state.get("logged_in", False):
    # 社員マスタから一致行を拾って自動ログイン
    _auto = df_login[df_login["社員ID"] == uid_q]
    if not _auto.empty and uid_q != "admin":
        st.session_state.logged_in = True
        st.session_state.user_id   = _auto.iloc[0]["社員ID"]
        st.session_state.user_name = _auto.iloc[0]["氏名"]
        st.session_state.dept      = _auto.iloc[0].get("部署", "") or ""
        st.session_state.is_admin  = False
        # 自動ログイン後にそのまま続行（rerunは不要）

# === クエリからのGPS取り込み（URLに gps / gps_error があればセッションへ反映） ===
qs = st.query_params
gps_q = qs.get("gps")
gps_err_q = qs.get("gps_error")

if gps_q or gps_err_q:
    # セッションキー初期化（無ければ）
    if "manual_gps" not in st.session_state:
        st.session_state.manual_gps = ""
    if "gps_error" not in st.session_state:
        st.session_state.gps_error = ""
    if "gps_click_token" not in st.session_state:
        st.session_state.gps_click_token = 0.0

    if gps_q:
        st.session_state.manual_gps = gps_q.strip()          # "lat,lng"
        st.session_state.gps_error = ""
    elif gps_err_q:
        st.session_state.manual_gps = ""
        st.session_state.gps_error = gps_err_q

    # 次回ポップアップが再起動しないようにトークンをリセット
    st.session_state.gps_click_token = 0.0

    # URLをきれいに（uid等は残しつつ gps クエリだけ除去）
    new_qs = {k: v for k, v in qs.items() if k not in ("gps", "gps_error")}
    st.query_params.clear()
    if new_qs:
        st.query_params.update(new_qs)

# ==============================
# ページネーション
# ==============================
def paginate_df(df: pd.DataFrame, page_key: str, per_page: int = 20):
    """
    DataFrame をページ分割して返す簡易ページネーション。
    - page_key：ページを保持する session_state のキー（ユニークに）
    - per_page：1ページの件数
    戻り値：表示用DF, 現在ページ番号, 最大ページ数
    """
    total = len(df)
    if total == 0:
        st.session_state[page_key] = 1
        st.caption("0件")
        return df, 1, 1

    max_page = max(1, math.ceil(total / per_page))
    cur = int(st.session_state.get(page_key, 1))
    cur = max(1, min(cur, max_page))

    colp1, colp2, colp3 = st.columns([1, 2, 1])
    with colp1:
        if st.button("◀ 前へ", disabled=(cur <= 1), key=f"{page_key}_prev"):
            cur = max(1, cur - 1)
    with colp2:
        cur = st.number_input("ページ", min_value=1, max_value=max_page, value=cur, step=1, key=f"{page_key}_num")
    with colp3:
        if st.button("次へ ▶", disabled=(cur >= max_page), key=f"{page_key}_next"):
            cur = min(max_page, cur + 1)

    st.session_state[page_key] = cur
    start = (cur - 1) * per_page
    end   = start + per_page
    st.caption(f"{total}件中 {start+1}–{min(end, total)} 件を表示（{cur}/{max_page}ページ）")
    return df.iloc[start:end], cur, max_page

# ==============================
# セッション初期化 & ログイン
# ==============================
if "logged_in" not in st.session_state:
    st.session_state.logged_in = False
    st.session_state.user_id   = ""
    st.session_state.user_name = ""
    st.session_state.dept      = ""
    st.session_state.is_admin  = False

if not st.session_state.logged_in:
    st.title("🔐 出退勤アプリ（ログイン式）")
    user_id = st.text_input("社員ID", max_chars=20)

    # admin のときだけパスワード欄を表示
    admin_pw = st.text_input("パスワード（adminのみ）", type="password") if user_id.strip() == "admin" else ""

    if st.button("ログイン"):
        if user_id.strip() == "admin":
            # ▼ admin を複数行許容：パスワードが一致する行を探す
            admins = df_login[df_login["社員ID"].astype(str).str.strip() == "admin"].copy()
            if admins.empty:
                st.error("管理者アカウントが見つかりません（社員ログイン情報.csv を確認してください）")
                st.stop()

            match = admins[admins["パスワード"].fillna("").astype(str).str.strip() == (admin_pw or "").strip()]
            if match.empty:
                st.error("管理者パスワードが正しくありません")
                st.stop()

            row = match.iloc[0]  # パス一致した行の氏名・部署を採用
            st.session_state.logged_in = True
            st.session_state.user_id   = "admin"
            st.session_state.user_name = row.get("氏名", "") or "管理者"
            st.session_state.dept      = row.get("部署", "") or ""
            st.session_state.is_admin  = True
            st.rerun()

        else:
            # ▼ 一般社員は従来どおり（社員ID一意想定）
            user = df_login[df_login["社員ID"] == user_id]
            if user.empty:
                st.error("ログイン失敗：社員IDが間違っています")
                st.stop()

            st.session_state.logged_in = True
            st.session_state.user_id   = user.iloc[0]["社員ID"]
            st.session_state.user_name = user.iloc[0]["氏名"]
            st.session_state.dept      = user.iloc[0].get("部署", "") or ""
            st.session_state.is_admin  = False

            st.query_params.update({"uid": st.session_state.user_id})

            st.rerun()

    st.stop()


# ここで is_admin を定義
is_admin = st.session_state.user_id == "admin"

# ==============================
# ログイン後の表示（サイドバー）
# ==============================
st.sidebar.markdown(f"👤 {st.session_state.user_name} さんがログイン中")
if st.session_state.dept:
    st.sidebar.caption(f"🏷 部署：{st.session_state.dept}")

if st.sidebar.button("ログアウト"):
    # セッション全消し
    for key in list(st.session_state.keys()):
        del st.session_state[key]

    # URLクエリから uid / gps / gps_error を除去（= 自動ログインを無効化）
    qs = dict(st.query_params)
    new_qs = {k: v for k, v in qs.items() if k not in ("uid", "gps", "gps_error")}
    st.query_params.clear()
    if new_qs:
        st.query_params.update(new_qs)

    st.rerun()

st.title("🕒 出退勤管理アプリ")

# ==============================
# 月選択（26日〜翌25日の締め）
# ==============================
st.subheader("📆 集計対象月の選択")

def get_month_period(selected_month: int, today: date):
    """
    月度: 26日～翌月25日
    selected_month は「締めの月」（例: 1=12/26～1/25）
    26日を起点に “今がどの締め月シーズンか” を判断する
    """
    # 26日を越えたら次月を“現在の締め月”として扱うアンカー
    anchor_year = today.year
    anchor_month = today.month + (1 if today.day >= 26 else 0)
    if anchor_month > 12:
        anchor_month -= 12
        anchor_year += 1

    base_year = anchor_year
    # 選択月がアンカー月より大きければ前年
    if selected_month > anchor_month:
        base_year -= 1

    if selected_month == 1:
        start = pd.to_datetime(f"{base_year-1}-12-26")
        end   = pd.to_datetime(f"{base_year}-01-25")
    else:
        start = pd.to_datetime(f"{base_year}-{selected_month-1:02d}-26")
        end   = pd.to_datetime(f"{base_year}-{selected_month:02d}-25")
    return start, end

# デフォルトのラジオ選択も“26日起点”で現在の締め月に合わせる
_today = today_jst()
_anchor_m = _today.month + (1 if _today.day >= 26 else 0)
if _anchor_m > 12:
    _anchor_m -= 12
default_idx = _anchor_m - 1  # 0〜11
selected_month = st.radio("📅 月を選択", list(range(1, 13)), index=default_idx, horizontal=True)
start_date, end_date = get_month_period(selected_month, _today)

st.caption(f"📅 表示期間：{start_date.strftime('%Y/%m/%d')} ～ {end_date.strftime('%Y/%m/%d')}")

def get_open_period(today_d: date):
    """今日が属する締め期間（26〜翌25日）を返す"""
    t = pd.Timestamp(today_d)
    for m in range(1, 13):
        s, e = get_month_period(m, today_d)
        if s <= t <= e:
            return s, e
    # 念のためのフォールバック
    return get_month_period(today_d.month, today_d)

OPEN_START, OPEN_END = get_open_period(today_jst())

# 勤怠データ前処理の直前あたりに差し込み
df_login_nodup = (
    df_login[df_login["社員ID"].astype(str).str.strip() != "admin"]
    .drop_duplicates(subset=["社員ID"], keep="first")
)
df_login_for_merge = pd.concat([
    df_login_nodup,
    df_login[df_login["社員ID"].astype(str).str.strip() == "admin"]  # 念のため残す
], ignore_index=True)

df = _read_csv_flexible(CSV_PATH).fillna("")
df = df.merge(df_login_for_merge[["社員ID", "部署"]], on="社員ID", how="left")

# ==============================
# 勤怠データ前処理
# ==============================
df["日付"] = pd.to_datetime(df["日付"], errors="coerce")
df["_出"]  = pd.to_datetime(df["出勤時刻"], format="%H:%M", errors="coerce")
df["_退"]  = pd.to_datetime(df["退勤時刻"], format="%H:%M", errors="coerce")

base_date = datetime.now(JST).replace(hour=0, minute=0, second=0, microsecond=0)
def _combine(t):
    return pd.Timestamp.combine(base_date.date(), t.time()) if pd.notna(t) else pd.NaT

df["出_dt"] = df["_出"].apply(_combine)
df["退_dt"] = df["_退"].apply(_combine)

fix_start = pd.Timestamp.combine(base_date.date(), pd.to_datetime("07:30").time())
fix_end   = pd.Timestamp.combine(base_date.date(), pd.to_datetime("17:00").time())

def calc_work_overtime(row):
    if pd.isna(row["出_dt"]) or pd.isna(row["退_dt"]) or row["退_dt"] < row["出_dt"]:
        return 0.0, 0.0
    dur_hours = (row["退_dt"] - row["出_dt"]).total_seconds() / 3600.0
    dept = (row.get("部署") or "").strip()
    if dept == "リサイクル事業部":
        before = max(0.0, (fix_start - row["出_dt"]).total_seconds()/3600.0) if row["出_dt"] < fix_start else 0.0
        after  = max(0.0, (row["退_dt"] - fix_end).total_seconds()/3600.0)  if row["退_dt"] > fix_end   else 0.0
        overtime = round(before + after, 2)
        work = round(dur_hours, 2)
        return work, overtime
    else:
        work_eff = max(0.0, dur_hours - 1.0)   # 休憩1h差引き
        overtime = max(0.0, work_eff - 8.0)    # 実働8h超
        return round(work_eff, 2), round(overtime, 2)

def format_hours_minutes(hours_float):
    total_minutes = int(round(float(hours_float) * 60)) if pd.notna(hours_float) else 0
    h, m = divmod(total_minutes, 60)
    if h and m:   return f"{h}時間{m}分"
    if h:         return f"{h}時間"
    if m:         return f"{m}分"
    return "0分"

def _is_hhmm(s: str) -> bool:
    return bool(re.fullmatch(r"([01]\d|2[0-3]):[0-5]\d", str(s).strip()))

# 勤怠データ前処理の後あたり
if df.empty:
    # 空でも列を用意しておく（float型で0行）
    df["勤務時間"] = pd.Series(dtype=float)
    df["残業時間"] = pd.Series(dtype=float)
else:
    results = df.apply(
        lambda r: pd.Series(calc_work_overtime(r), index=["勤務時間", "残業時間"]),
        axis=1
    )
    df[["勤務時間", "残業時間"]] = results

# 以降はそのままでOK
df["勤務時間"] = df["勤務時間"].fillna(0).astype(float).round(2)
df["残業時間"] = df["残業時間"].fillna(0).astype(float).round(2)

# ==============================
# 承認済み残業計算
# ==============================
def apply_approved_overtime(df_att: pd.DataFrame) -> pd.DataFrame:
    """承認済みの残業申請(OVERTIME_CSV)で、残業時間を上書きする"""
    ot = read_overtime_csv()
    if ot.empty:
        df_att["承認残業時間"] = df_att["残業時間"].astype(float)
        return df_att
    # 承認のみ抽出
    ok = ot[ot["ステータス"] == "承認"].copy()
    if ok.empty:
        df_att["承認残業時間"] = df_att["残業時間"].astype(float)
        return df_att
    # 型合わせ
    df2 = df_att.copy()
    df2["日付_str"] = df2["日付"].dt.strftime("%Y-%m-%d")
    # key: 社員ID+日付 でマージ
    df2 = df2.merge(
        ok[["社員ID","対象日","申請残業H"]].rename(columns={"対象日":"日付_str"}),
        on=["社員ID","日付_str"], how="left"
    )
    # 申請残業H があればそれを優先、無ければ元の残業時間
    def _pick(row):
        try:
            v = float(str(row.get("申請残業H","")).strip()) if str(row.get("申請残業H","")).strip() else None
        except:
            v = None
        # 承認レコードがなければ自動計算値のまま
        return v if v is not None else float(row["残業時間"])
    df2["承認残業時間"] = df2.apply(_pick, axis=1).astype(float).round(2)
    return df2.drop(columns=["申請残業H"])

# 実行
df = apply_approved_overtime(df)


# ==============================
# 分岐：管理者 or 社員
# ==============================
# ==============================
# 管理者UI（サイドバー切替付き）
# ==============================
if is_admin:
    # ▼ 管理者用サイドバー切替
    admin_menu = st.sidebar.radio(
        "📑 管理メニュー",
        ["各自の出退勤確認", "申請（承認/却下）", "ダウンロード・保守"],
        index=0,
        key="admin_main_view_selector"
    )

    # ---------------------------------
    # A) 各自の出退勤確認
    # ---------------------------------
    if admin_menu == "各自の出退勤確認":
        st.header("👥 各自の出退勤確認")

        # 社員選択（admin を除外）
        all_users = (
            df_login[df_login["社員ID"].astype(str).str.strip() != "admin"][["社員ID", "氏名"]]
            .drop_duplicates()
            .copy()
        )

        if all_users.empty:
            st.warning("社員マスタに表示可能な社員がいません（adminのみの可能性）。")
            st.stop()

        all_users["表示名"] = all_users["社員ID"].astype(str).str.strip() + "：" + all_users["氏名"].astype(str).str.strip()
        selected_label = st.selectbox("社員を選択して出退勤履歴を表示", all_users["表示名"])
        selected_user_id = selected_label.split("：", 1)[0]
        selected_user_name = all_users.loc[
            all_users["社員ID"].astype(str).str.strip() == selected_user_id, "氏名"
        ].values[0]

        # 期間＆対象社員で絞り込み
        df_admin_user = df[
            (df["社員ID"] == selected_user_id) &
            (df["日付"] >= start_date) &
            (df["日付"] <= end_date)
        ].sort_values("日付")

        if df_admin_user.empty:
            st.info(f"{selected_user_name} さんのこの月の出退勤記録はありません。")
        else:
            # 表示整形
            df_show = df_admin_user.copy()
            df_show["日付"] = df_show["日付"].dt.strftime("%Y-%m-%d")
            df_show = df_show.rename(columns={
                "日付": "日付", "出勤時刻": "出勤", "退勤時刻": "退勤",
                "勤務時間": "勤務H", "残業時間": "残業H"
            })
            df_show["勤務H"] = df_show["勤務H"].astype(float).apply(format_hours_minutes)
            df_show["残業H"] = df_show["残業H"].astype(float).apply(format_hours_minutes)
            df_show["残業H(承認)"] = df_show["承認残業時間"].astype(float).apply(format_hours_minutes)

            st.dataframe(
                df_show[["日付", "出勤", "退勤", "勤務H", "残業H", "残業H(承認)"]],
                hide_index=True,
                use_container_width=True
            )

            # 位置情報
            gps_df = (
                df_admin_user[["日付", "緯度", "経度"]].copy()
                if {"緯度", "経度"}.issubset(df_admin_user.columns) else
                pd.DataFrame(columns=["日付", "緯度", "経度"])
            )
            if not gps_df.empty:
                gps_df["日付"] = gps_df["日付"].dt.strftime("%Y-%m-%d")

            with st.expander(f"📍 位置情報（{selected_user_name} さん）", expanded=False):
                if not gps_df.empty:
                    links_df = gps_df.copy()
                    links_df["GoogleMap"] = links_df.apply(
                        lambda r: (
                            f"https://www.google.com/maps?q={r['緯度']},{r['経度']}"
                            if (str(r.get("緯度", "")).strip() and str(r.get("経度", "")).strip()) else ""
                        ),
                        axis=1
                    )
                    links_df = links_df[["日付", "GoogleMap"]]
                    try:
                        st.dataframe(
                            links_df, hide_index=True, use_container_width=True,
                            column_config={
                                "日付": st.column_config.TextColumn("日付"),
                                "GoogleMap": st.column_config.LinkColumn("地図で見る", display_text="地図で見る"),
                            }
                        )
                    except Exception:
                        st.dataframe(
                            links_df.rename(columns={"GoogleMap": "地図URL"}),
                            hide_index=True, use_container_width=True
                        )
                else:
                    st.caption("位置情報はありません。")

            with st.expander(f"📄 出退勤履歴（{selected_user_name} さん）", expanded=False):
                st.dataframe(
                    df_show[["日付", "出勤", "退勤", "勤務H", "残業H"]],
                    hide_index=True,
                    use_container_width=True
                )
                total_ot_calc = float(df_admin_user["残業時間"].sum())
                total_ot_approved = float(df_admin_user["承認残業時間"].sum())
                st.subheader(f"⏱️ 合計残業時間（自動計算）：{format_hours_minutes(total_ot_calc)}")
                st.subheader(f"✅ 合計残業時間（承認反映）：{format_hours_minutes(total_ot_approved)}")

            # ===== 修正 =====
            with st.expander(f"✏️ 出退勤の修正（{selected_user_name} さん）", expanded=False):
                edit_df = df_admin_user[["日付", "出勤時刻", "退勤時刻"]].copy().sort_values("日付")
                edit_df["日付"] = edit_df["日付"].dt.strftime("%Y-%m-%d")
                edit_df = edit_df.reset_index(drop=True)

                edited = st.data_editor(
                    edit_df,
                    use_container_width=True,
                    hide_index=True,
                    num_rows="fixed",
                    column_config={
                        "日付": st.column_config.TextColumn("日付", disabled=True),
                        "出勤時刻": st.column_config.TextColumn("出勤時刻（HH:MM）"),
                        "退勤時刻": st.column_config.TextColumn("退勤時刻（HH:MM）"),
                    },
                    key="admin_edit_editor",
                )

                if st.button("💾 修正内容を保存", type="primary", key="admin_save_edits"):
                    base = _read_csv_flexible(CSV_PATH).fillna("")
                    errors = []; updated = False
                    for _, r in edited.iterrows():
                        d  = str(r["日付"])
                        sh = str(r["出勤時刻"]).strip()
                        eh = str(r["退勤時刻"]).strip()
                        row_errs = []
                        if sh and not _is_hhmm(sh): row_errs.append(f"{d} の出勤時刻が不正: {sh}")
                        if eh and not _is_hhmm(eh): row_errs.append(f"{d} の退勤時刻が不正: {eh}")
                        if row_errs:
                            errors.extend(row_errs)
                            continue
                        m = (base["社員ID"] == selected_user_id) & (base["日付"] == d)
                        if not m.any():
                            base = pd.concat([base, pd.DataFrame([{
                                "社員ID": selected_user_id, "氏名": selected_user_name,
                                "日付": d, "出勤時刻": sh, "退勤時刻": eh,
                            }])], ignore_index=True)
                        else:
                            if sh != "": base.loc[m, "出勤時刻"] = sh
                            if eh != "": base.loc[m, "退勤時刻"] = eh
                        updated = True

                    if updated and safe_write_csv(base, CSV_PATH, ATT_COLUMNS):
                        st.success("正常な行は保存しました。最新表示に更新します。")
                        time.sleep(1.0); st.rerun()
                    if errors:
                        st.warning("以下の行は保存できませんでした：\n- " + "\n- ".join(errors))

            # ===== 削除 =====
            with st.expander(f"🗑️ 出退勤の削除（{selected_user_name} さん）", expanded=False):
                del_df = df_admin_user[["日付", "出勤時刻", "退勤時刻"]].copy().sort_values("日付")
                del_df["日付"] = del_df["日付"].dt.strftime("%Y-%m-%d")
                del_df = del_df.reset_index(drop=True).assign(削除=False)

                edited_del = st.data_editor(
                    del_df,
                    use_container_width=True,
                    hide_index=True,
                    num_rows="fixed",
                    column_config={
                        "削除": st.column_config.CheckboxColumn("削除", help="削除する行にチェック"),
                        "日付": st.column_config.TextColumn("日付", disabled=True),
                        "出勤時刻": st.column_config.TextColumn("出勤時刻", disabled=True),
                        "退勤時刻": st.column_config.TextColumn("退勤時刻", disabled=True),
                    },
                    key="admin_delete_editor",
                )
                to_delete = edited_del[edited_del["削除"] == True]["日付"].tolist()
                col_a, col_b = st.columns([1, 2])
                with col_a:
                    confirm = st.checkbox("本当に削除します", key="admin_delete_confirm")
                with col_b:
                    if st.button("❌ チェックした行を削除", disabled=(len(to_delete) == 0 or not confirm),
                                 key="admin_delete_button"):
                        base = _read_csv_flexible(CSV_PATH).fillna("")
                        before = len(base)
                        mask = (base["社員ID"] == selected_user_id) & (base["日付"].isin(to_delete))
                        base = base[~mask]
                        removed = before - len(base)
                        if safe_write_csv(base, CSV_PATH, ATT_COLUMNS):
                            st.success(f"{removed} 行を削除しました。最新表示に更新します。")
                            time.sleep(1.0); st.rerun()

    # ---------------------------------
    # B) 申請（承認/却下）
    # ---------------------------------
    elif admin_menu == "申請（承認/却下）":
        st.header("✅ 申請（承認/却下）")

        # --- 残業申請の承認／却下 ---
        with st.expander("⏱️ 残業申請の承認／却下", expanded=False):
            ot = read_overtime_csv().merge(df_login[["社員ID","部署"]], on="社員ID", how="left")
            start_s = start_date.strftime("%Y-%m-%d"); end_s = end_date.strftime("%Y-%m-%d")
            mask_period = (ot["対象日"] >= start_s) & (ot["対象日"] <= end_s)

            col1, col2, col3 = st.columns([2, 2, 1.4])
            with col1:
                status_filter_ot = st.multiselect(
                    "対象ステータス", ["申請済", "承認", "却下"],
                    default=["申請済"], key="admin_overtime_status_filter"
                )
            with col2:
                dept_options_ot = sorted([d for d in ot["部署"].dropna().unique().tolist() if str(d).strip()])
                dept_filter_ot = st.multiselect(
                    "部署で絞り込み", dept_options_ot, default=[], key="admin_overtime_dept_filter"
                )
            with col3:
                st.caption(f"期間: {start_s} ～ {end_s}")

            m = mask_period
            if status_filter_ot: m &= ot["ステータス"].isin(status_filter_ot)
            if dept_filter_ot:   m &= ot["部署"].isin(dept_filter_ot)

            # ▼ これまで通り抽出
            ot_view = ot.loc[m, [
                "社員ID","氏名","部署","対象日","申請日時","申請残業H","申請理由",
                "ステータス","承認者","承認日時","却下理由"
            ]].copy().sort_values(["ステータス","対象日","社員ID"])

            if ot_view.empty:
                st.caption("この条件に該当する申請はありません。")
            else:
                # 小数時間→分表示（UI用）
                def _h_to_min_text(x):
                    try:
                        return f"{int(round(float(str(x).strip() or 0) * 60))}分"
                    except Exception:
                        return ""
                ot_view["申請残業(分)"] = ot_view["申請残業H"].apply(_h_to_min_text)

                # 表示列だけに絞る（キー3列は必ず残す）
                ot_view = ot_view[[
                    "社員ID","氏名","部署","対象日","申請日時","申請残業(分)","申請理由",
                    "ステータス","承認者","承認日時","却下理由"
                ]].copy()

                # 操作用のチェック列を付与（このDFに直接追加）
                for _c in ["承認","却下","承認解除","削除"]:
                    ot_view[_c] = False
                ot_view["却下理由(入力)"] = ""

                edited = st.data_editor(
                    ot_view, hide_index=True, use_container_width=True,
                    column_config={
                        "社員ID": st.column_config.TextColumn("社員ID", disabled=True),
                        "氏名": st.column_config.TextColumn("氏名", disabled=True),
                        "部署": st.column_config.TextColumn("部署", disabled=True),
                        "対象日": st.column_config.TextColumn("対象日", disabled=True),
                        "申請日時": st.column_config.TextColumn("申請日時", disabled=True),
                        "申請残業(分)": st.column_config.TextColumn("申請残業（分表示）", disabled=True),
                        "申請理由": st.column_config.TextColumn("申請理由", disabled=True),
                        "ステータス": st.column_config.TextColumn("現ステータス", disabled=True),
                        "承認者": st.column_config.TextColumn("承認者", disabled=True),
                        "承認日時": st.column_config.TextColumn("承認日時", disabled=True),
                        "却下理由": st.column_config.TextColumn("却下理由(既存)", disabled=True),
                        "承認": st.column_config.CheckboxColumn("承認する"),
                        "却下": st.column_config.CheckboxColumn("却下する"),
                        "承認解除": st.column_config.CheckboxColumn("承認を取り消す"),
                        "削除": st.column_config.CheckboxColumn("削除（申請済のみ）"),
                        "却下理由(入力)": st.column_config.TextColumn("却下理由（入力）"),
                    },
                    key="overtime_approvals_editor"
                )

                colb1, colb2 = st.columns([1, 3])
                with colb1:
                    apply_clicked = st.button("💾 選択を反映", type="primary", key="ot_apply")
                with colb2:
                    st.caption("※ 同じ行で複数操作は不可。却下時は理由を入力。")

                if apply_clicked:
                    approver = st.session_state.user_name or "admin"
                    when_ts = datetime.now(JST).strftime("%Y-%m-%d %H:%M:%S")
                    base = read_overtime_csv()
                    applied = 0; conflicts = []; logs = []

                    for _, r in edited.iterrows():
                        approve = bool(r.get("承認", False))
                        reject  = bool(r.get("却下", False))
                        unapp   = bool(r.get("承認解除", False))
                        delete  = bool(r.get("削除", False))
                        if sum([approve, reject, unapp, delete]) == 0:
                            continue
                        if sum([approve, reject, unapp, delete]) > 1:
                            conflicts.append(f'{r["氏名"]} {r["対象日"]}: 同時に複数操作はできません')
                            continue

                        km = (
                            (base["社員ID"] == r["社員ID"]) &
                            (base["対象日"] == r["対象日"]) &
                            (base["申請日時"] == r["申請日時"])
                        )
                        if not km.any():
                            conflicts.append(f'{r["氏名"]} {r["対象日"]}: 対象が見つかりません')
                            continue

                        cur = str(base.loc[km, "ステータス"].iloc[0])
                        if approve:
                            if cur != "申請済":
                                conflicts.append(f'{r["氏名"]} {r["対象日"]}: 現在 {cur} で承認不可')
                                continue
                            base.loc[km, ["ステータス","承認者","承認日時","却下理由"]] = ["承認", approver, when_ts, ""]
                            new_status = "承認"
                        elif reject:
                            if cur != "申請済":
                                conflicts.append(f'{r["氏名"]} {r["対象日"]}: 現在 {cur} で却下不可')
                                continue
                            rsn = str(r.get("却下理由(入力)", "")).strip()
                            if not rsn:
                                conflicts.append(f'{r["氏名"]} {r["対象日"]}: 却下理由が未入力')
                                continue
                            base.loc[km, ["ステータス","承認者","承認日時","却下理由"]] = ["却下", approver, when_ts, rsn]
                            new_status = "却下"
                        elif unapp:
                            if cur != "承認":
                                conflicts.append(f'{r["氏名"]} {r["対象日"]}: 現在 {cur} で承認解除不可')
                                continue
                            base.loc[km, ["ステータス","承認者","承認日時","却下理由"]] = ["申請済", "", "", ""]
                            new_status = "申請済"
                        else:
                            if cur != "申請済":
                                conflicts.append(f'{r["氏名"]} {r["対象日"]}: 現在 {cur} で削除不可（申請済のみ）')
                                continue
                            base = base.loc[~km].copy()
                            new_status = "申請削除"

                        applied += int(km.sum())
                        logs.append({
                            "timestamp": when_ts, "承認者": approver,
                            "社員ID": r["社員ID"], "氏名": r["氏名"],
                            "休暇日": r["対象日"], "申請日": r["申請日時"],
                            "旧ステータス": cur, "新ステータス": f"残業:{new_status}",
                            "却下理由": r.get("却下理由(入力)", "")
                        })

                    if applied > 0:
                        write_overtime_csv(base)
                        append_audit_log(logs)
                        st.success(f"{applied} 件を更新しました。")
                        time.sleep(1); st.rerun()
                    if conflicts:
                        st.warning("一部適用できませんでした：\n- " + "\n- ".join(conflicts))

        # --- 休日申請の承認／却下 ---
        with st.expander("📅 休日申請の承認／却下", expanded=False):
            hd = read_holiday_csv().merge(df_login[["社員ID", "部署"]], on="社員ID", how="left")
            start_s = start_date.strftime("%Y-%m-%d"); end_s = end_date.strftime("%Y-%m-%d")
            period_mask = (hd["休暇日"] >= start_s) & (hd["休暇日"] <= end_s)

            col1, col2, col3 = st.columns([2, 2, 1.4])
            with col1:
                status_filter_hd = st.multiselect(
                    "対象ステータス", ["申請済", "承認", "却下"],
                    default=["申請済"], key="admin_holiday_status_filter"
                )
            with col2:
                dept_options_hd = sorted([d for d in hd["部署"].dropna().unique().tolist() if str(d).strip()])
                dept_filter_hd = st.multiselect(
                    "部署で絞り込み", dept_options_hd, default=[], key="admin_holiday_dept_filter"
                )
            with col3:
                st.caption(f"期間: {start_s} ～ {end_s}")

            mask = period_mask
            if status_filter_hd: mask &= hd["ステータス"].isin(status_filter_hd)
            if dept_filter_hd:   mask &= hd["部署"].isin(dept_filter_hd)

            hd_view = hd.loc[mask, [
                "社員ID","氏名","部署","申請日","休暇日","休暇種類","備考",
                "ステータス","承認者","承認日時","却下理由"
            ]].copy().sort_values(["ステータス","休暇日","社員ID"])

            if hd_view.empty:
                st.caption("この条件に該当する申請はありません。")
            else:
                hd_view["承認"] = False
                hd_view["却下"] = False
                hd_view["却下理由(入力)"] = ""
                hd_view["承認解除"] = False
                hd_view["削除"] = False

                edited = st.data_editor(
                    hd_view, hide_index=True, use_container_width=True,
                    column_config={
                        "社員ID": st.column_config.TextColumn("社員ID", disabled=True),
                        "氏名": st.column_config.TextColumn("氏名", disabled=True),
                        "部署": st.column_config.TextColumn("部署", disabled=True),
                        "申請日": st.column_config.TextColumn("申請日", disabled=True),
                        "休暇日": st.column_config.TextColumn("休暇日", disabled=True),
                        "休暇種類": st.column_config.TextColumn("休暇種類", disabled=True),
                        "備考": st.column_config.TextColumn("備考", disabled=True),
                        "ステータス": st.column_config.TextColumn("現ステータス", disabled=True),
                        "承認者": st.column_config.TextColumn("承認者", disabled=True),
                        "承認日時": st.column_config.TextColumn("承認日時", disabled=True),
                        "却下理由": st.column_config.TextColumn("却下理由(既存)", disabled=True),
                        "承認": st.column_config.CheckboxColumn("承認する"),
                        "却下": st.column_config.CheckboxColumn("却下する"),
                        "却下理由(入力)": st.column_config.TextColumn("却下理由（入力）"),
                        "承認解除": st.column_config.CheckboxColumn("承認を取り消す"),
                        "削除": st.column_config.CheckboxColumn("削除（申請済のみ）"),
                    },
                    key="holiday_approvals_editor"
                )

                colb1, colb2 = st.columns([1, 3])
                with colb1:
                    apply_clicked = st.button("💾 選択を反映", type="primary")
                with colb2:
                    st.caption("※ 同じ行で「承認」と「却下」を同時に選ばないでください。却下時は理由を入力。")

                if apply_clicked:
                    approver = st.session_state.user_name or "admin"
                    when_ts = datetime.now(JST).strftime("%Y-%m-%d %H:%M:%S")

                    base = read_holiday_csv()
                    to_change = []
                    conflicts = []

                    for _, r in edited.iterrows():
                        approve   = bool(r.get("承認", False))
                        reject    = bool(r.get("却下", False))
                        unapprove = bool(r.get("承認解除", False))
                        delete_it = bool(r.get("削除", False))

                        if sum([approve, reject, unapprove, delete_it]) == 0:
                            continue
                        if sum([approve, reject, unapprove, delete_it]) > 1:
                            conflicts.append(f'{r["氏名"]} {r["休暇日"]}: 承認/却下/承認解除/削除は同時に選べません')
                            continue

                        key_mask = (
                            (base["社員ID"] == r["社員ID"]) &
                            (base["休暇日"] == r["休暇日"]) &
                            (base["申請日"] == r["申請日"])
                        )
                        if not key_mask.any():
                            conflicts.append(f'{r["氏名"]} {r["休暇日"]}: 対象レコードが見つかりません')
                            continue

                        cur_status = str(base.loc[key_mask, "ステータス"].iloc[0])

                        if approve:
                            if cur_status != "申請済":
                                conflicts.append(f'{r["氏名"]} {r["休暇日"]}: 現在 {cur_status} のため承認できません')
                                continue
                            action, reason = "承認", ""
                        elif reject:
                            if cur_status != "申請済":
                                conflicts.append(f'{r["氏名"]} {r["休暇日"]}: 現在 {cur_status} のため却下できません')
                                continue
                            reason = str(r.get("却下理由(入力)", "")).strip()
                            if not reason:
                                conflicts.append(f'{r["氏名"]} {r["休暇日"]}: 却下理由が未入力')
                                continue
                            action = "却下"
                        elif unapprove:
                            if cur_status != "承認":
                                conflicts.append(f'{r["氏名"]} {r["休暇日"]}: 現在 {cur_status} のため承認解除できません')
                                continue
                            action, reason = "承認解除", ""
                        else:
                            if cur_status != "申請済":
                                conflicts.append(f'{r["氏名"]} {r["休暇日"]}: 現在 {cur_status} のため削除できません')
                                continue
                            action, reason = "削除", ""

                        to_change.append({
                            "社員ID": r["社員ID"], "氏名": r["氏名"],
                            "休暇日": r["休暇日"], "申請日": r["申請日"],
                            "action": action, "reason": reason, "old_status": cur_status
                        })

                    if not to_change and not conflicts:
                        st.info("変更はありません。")
                    else:
                        latest = read_holiday_csv()
                        applied = 0
                        audit_rows = []

                        for ch in to_change:
                            km = (
                                (latest["社員ID"] == ch["社員ID"]) &
                                (latest["休暇日"] == ch["休暇日"]) &
                                (latest["申請日"] == ch["申請日"])
                            )
                            if not km.any():
                                conflicts.append(f'{ch["氏名"]} {ch["休暇日"]}: 直前に削除/変更され見つかりません')
                                continue

                            cur2 = str(latest.loc[km, "ステータス"].iloc[0])
                            if ch["action"] in ("承認", "却下") and cur2 != "申請済":
                                conflicts.append(f'{ch["氏名"]} {ch["休暇日"]}: 直前に {cur2} に更新されスキップ')
                                continue
                            if ch["action"] == "承認解除" and cur2 != "承認":
                                conflicts.append(f'{ch["氏名"]} {ch["休暇日"]}: 直前に {cur2} に更新されスキップ')
                                continue

                            if ch["action"] == "承認":
                                latest.loc[km, ["ステータス","承認者","承認日時","却下理由"]] = ["承認", approver, when_ts, ""]
                                new_status_for_audit = "承認"
                            elif ch["action"] == "却下":
                                latest.loc[km, ["ステータス","承認者","承認日時","却下理由"]] = ["却下", approver, when_ts, ch["reason"]]
                                new_status_for_audit = "却下"
                            elif ch["action"] == "承認解除":
                                latest.loc[km, ["ステータス","承認者","承認日時","却下理由"]] = ["申請済", "", "", ""]
                                new_status_for_audit = "申請済"
                            else:  # 削除
                                latest = latest.loc[~km].copy()
                                new_status_for_audit = "申請削除"

                            applied += int(km.sum())
                            audit_rows.append({
                                "timestamp": when_ts, "承認者": approver,
                                "社員ID": ch["社員ID"], "氏名": ch["氏名"],
                                "休暇日": ch["休暇日"], "申請日": ch["申請日"],
                                "旧ステータス": ch["old_status"], "新ステータス": new_status_for_audit,
                                "却下理由": ch["reason"],
                            })

                        if applied > 0:
                            write_holiday_csv(latest)
                            append_audit_log(audit_rows)
                            st.success(f"{applied} 件を更新しました。")

                        if conflicts:
                            st.warning("一部の行は適用できませんでした：\n- " + "\n- ".join(conflicts))

                        if applied > 0:
                            time.sleep(1.0); st.rerun()

        # --- 監査ログ ---
        with st.expander("📝 監査ログ（承認/却下の履歴）", expanded=False):
            if os.path.exists(AUDIT_LOG_CSV):
                try:
                    log_df = pd.read_csv(AUDIT_LOG_CSV, dtype=str, encoding="utf-8-sig").fillna("")
                except UnicodeDecodeError:
                    log_df = pd.read_csv(AUDIT_LOG_CSV, dtype=str, encoding="cp932", encoding_errors="replace").fillna("")
            else:
                log_df = pd.DataFrame(columns=AUDIT_COLUMNS)

            if log_df.empty:
                st.caption("監査ログはまだありません。")
            else:
                start_s = start_date.strftime("%Y-%m-%d"); end_s = end_date.strftime("%Y-%m-%d")
                col1, col2, col3 = st.columns([1.4, 1.4, 2])
                with col1:
                    date_from = st.text_input("開始日 (YYYY-MM-DD)", value=start_s)
                with col2:
                    date_to   = st.text_input("終了日 (YYYY-MM-DD)", value=end_s)
                with col3:
                    approver = st.text_input("承認者で絞り込み（任意）", value="")

                dfv = log_df.copy()
                if date_from: dfv = dfv[dfv["timestamp"].str[:10] >= date_from]
                if date_to:   dfv = dfv[dfv["timestamp"].str[:10] <= date_to]
                if approver.strip(): dfv = dfv[dfv["承認者"].str.contains(approver.strip(), na=False)]

                show = dfv[["timestamp","承認者","社員ID","氏名","休暇日","申請日","旧ステータス","新ステータス","却下理由"]]\
                       .sort_values(["timestamp"], ascending=False)
                st.dataframe(show, hide_index=True, use_container_width=True)

                xls_buf = io.BytesIO()
                with pd.ExcelWriter(xls_buf, engine="openpyxl") as writer:
                    show.to_excel(writer, index=False, sheet_name="監査ログ")
                st.download_button(
                    "⬇️ 監査ログをExcelでダウンロード",
                    data=xls_buf.getvalue(),
                    file_name=f"監査ログ_{start_s}_to_{end_s}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

    # ---------------------------------
    # C) ダウンロード・保守
    # ---------------------------------
    else:
        st.header("📦 ダウンロード・保守")

        # 文字化け修復
        with st.expander("🧹 文字化け修復（氏名を社員マスタで一括上書き）", expanded=False):
            st.caption("※ 初回運用で氏名の文字化けが発生した場合のみ使用してください。")
            if st.button("氏名を一括修復して保存"):
                base = _read_csv_flexible(CSV_PATH).fillna("")
                base = base.drop(columns=["氏名"], errors="ignore") \
                           .merge(df_login[["社員ID","氏名"]], on="社員ID", how="left")
                if safe_write_csv(base, CSV_PATH, ATT_COLUMNS):
                    st.success("氏名を社員マスタで上書きしました。")
                    time.sleep(1.0); st.rerun()

        # 全社員のエクスポート（勤務＋休日申請）
        with st.expander("📥 全社員のデータをダウンロード", expanded=False):
            export_df = df[(df["日付"] >= start_date) & (df["日付"] <= end_date)].copy()
            export_df = export_df.drop(columns=["氏名"], errors="ignore") \
                                 .merge(df_login[["社員ID", "氏名"]], on="社員ID", how="left")
            export_df["日付"] = export_df["日付"].dt.strftime("%Y-%m-%d")
            cols = ["社員ID","氏名","日付","出勤時刻","退勤時刻","勤務時間","残業時間","承認残業時間"]
            export_df = export_df.reindex(columns=[c for c in cols if c in export_df.columns])

            ym_name = f"{end_date.year}-{end_date.month:02d}"

            # 休日申請データ
            hd_all = read_holiday_csv().merge(df_login[["社員ID", "部署"]], on="社員ID", how="left")
            start_s = start_date.strftime("%Y-%m-%d"); end_s = end_date.strftime("%Y-%m-%d")
            mask = (hd_all["休暇日"] >= start_s) & (hd_all["休暇日"] <= end_s)
            hd_export = hd_all.loc[mask, [
                "社員ID","氏名","部署","申請日","休暇日","休暇種類","備考","ステータス","承認者","承認日時","却下理由"
            ]].copy()
            hd_export["申請日"]   = pd.to_datetime(hd_export["申請日"],   errors="coerce")
            hd_export["休暇日"]   = pd.to_datetime(hd_export["休暇日"],   errors="coerce")
            hd_export["承認日時"] = pd.to_datetime(hd_export["承認日時"], errors="coerce")
            hd_export = hd_export.sort_values(["休暇日", "社員ID"])

            xls_buf = io.BytesIO()
            with pd.ExcelWriter(xls_buf, engine="openpyxl") as writer:
                export_df.to_excel(writer, index=False, sheet_name="勤務実績")
                hd_export.to_excel(writer, index=False, sheet_name="休日申請")

                from openpyxl.utils import get_column_letter
                wb  = writer.book
                ws1 = writer.sheets["勤務実績"]
                ws2 = writer.sheets["休日申請"]

                def beautify(ws):
                    ws.auto_filter.ref = ws.dimensions
                    ws.freeze_panes = "A2"
                    for col_idx, col_cells in enumerate(ws.columns, start=1):
                        max_len = 0
                        for cell in col_cells:
                            val = "" if cell.value is None else str(cell.value)
                            max_len = max(max_len, len(val))
                        from openpyxl.utils import get_column_letter as gl
                        ws.column_dimensions[gl(col_idx)].width = min(max(max_len + 2, 8), 40)

                beautify(ws1); beautify(ws2)

                headers = [c.value for c in next(ws2.iter_rows(min_row=1, max_row=1))]
                def col_letter(col_name: str):
                    from openpyxl.utils import get_column_letter
                    idx = headers.index(col_name) + 1
                    return get_column_letter(idx), idx

                try:
                    if ws2.max_row >= 2:
                        col申請, _ = col_letter("申請日")
                        col休暇, _ = col_letter("休暇日")
                        col承認時, _ = col_letter("承認日時")
                        for row in range(2, ws2.max_row + 1):
                            ws2[f"{col申請}{row}"].number_format = "yyyy-mm-dd"
                            ws2[f"{col休暇}{row}"].number_format = "yyyy-mm-dd"
                            ws2[f"{col承認時}{row}"].number_format = "yyyy-mm-dd hh:mm"
                except ValueError:
                    pass

                try:
                    if ws2.max_row >= 2 and "ステータス" in headers:
                        from openpyxl.styles import PatternFill
                        from openpyxl.formatting.rule import CellIsRule
                        colステータス, _ = col_letter("ステータス")
                        status_range = f"{colステータス}2:{colステータス}{ws2.max_row}"
                        fill_pending  = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                        fill_approved = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        fill_rejected = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
                        ws2.conditional_formatting.add(status_range, CellIsRule(operator="equal", formula=['"申請済"'], stopIfTrue=False, fill=fill_pending))
                        ws2.conditional_formatting.add(status_range, CellIsRule(operator="equal", formula=['"承認"'], stopIfTrue=False, fill=fill_approved))
                        ws2.conditional_formatting.add(status_range, CellIsRule(operator="equal", formula=['"却下"'], stopIfTrue=False, fill=fill_rejected))
                except ValueError:
                    pass

            st.download_button(
                "⬇️ Excel(.xlsx)でダウンロード（勤務＋申請の2枚シート）",
                data=xls_buf.getvalue(),
                file_name=f"全社員_勤務実績_休日申請_{ym_name}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            try:
                csv_bytes = export_df.to_csv(index=False, encoding="cp932").encode("cp932", errors="replace")
            except Exception:
                csv_bytes = export_df.to_csv(index=False).encode("cp932", errors="replace")
            st.download_button(
                "⬇️ CSV(Shift_JIS/cp932)でダウンロード",
                data=csv_bytes,
                file_name=f"全社員_出退勤履歴_{ym_name}.csv",
                mime="text/csv",
            )

        # バックアップ／復元
        with st.expander("💾 バックアップ（ZIP）／🛠️ 復元（ZIP/CSV）", expanded=False):
            st.markdown("**推奨運用**：業務終了時に必ずZIPでバックアップ → ローカルPCに保管。")

            col_b1, col_b2 = st.columns([1.2, 2])
            with col_b1:
                buf = io.BytesIO()
                with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
                    for path, cols, fname in BACKUP_TABLES:
                        dfb = _read_existing_or_empty(path, cols)
                        content = dfb[cols].to_csv(index=False)
                        zf.writestr(fname, content.encode("cp932", errors="replace"))
                st.download_button(
                    "⬇️ 全CSVをZIPでダウンロード",
                    data=buf.getvalue(),
                    file_name=f"backup_{datetime.now():%Y%m%d_%H%M%S}.zip",
                    mime="application/zip",
                    use_container_width=True
                )
            with col_b2:
                st.caption("内容：attendance_log.csv / holiday_requests.csv / holiday_audit_log.csv / 社員ログイン情報.csv")

            st.markdown("---")

            uploads = st.file_uploader(
                "ZIP（4ファイルまとめ）または個別CSVを1つ以上アップロード",
                type=["zip", "csv"], accept_multiple_files=True
            )
            c1, c2 = st.columns([1.2, 2])
            with c1:
                do_backup = st.checkbox("上書き前に既存をZIPバックアップする", value=True)
            with c2:
                st.caption("※ 必須列が欠けたCSVはスキップされます。ZIPは上の4ファイル名で構成されている想定です。")

            if st.button("インポートを実行", type="primary", disabled=(not uploads)):
                if do_backup:
                    try:
                        buf = io.BytesIO()
                        with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
                            for path, cols, fname in BACKUP_TABLES:
                                dfb = _read_existing_or_empty(path, cols)
                                content = dfb[cols].to_csv(index=False)
                                zf.writestr(fname, content.encode("cp932"))
                        backup_dir = os.path.join(DATA_DIR, "backups")
                        os.makedirs(backup_dir, exist_ok=True)
                        backup_path = os.path.join(backup_dir, f"pre_import_{datetime.now():%Y%m%d_%H%M%S}.zip")
                        with open(backup_path, "wb") as f:
                            f.write(buf.getvalue())
                        st.info(f"既存データをバックアップしました：{backup_path}")
                    except Exception as e:
                        st.warning(f"バックアップで警告：{e}")

                incoming: dict[str, bytes] = {}
                for up in uploads:
                    name = (up.name or "").split("/")[-1]
                    if name.lower().endswith(".zip"):
                        try:
                            with zipfile.ZipFile(up) as zf:
                                for n in zf.namelist():
                                    if n.lower().endswith(".csv"):
                                        incoming[n.split("/")[-1]] = zf.read(n)
                        except Exception as e:
                            st.error(f"ZIPの解凍に失敗：{name} / {e}")
                    else:
                        incoming[name] = up.read()

                applied, skipped, errors = [], [], []
                for path, cols, fname in BACKUP_TABLES:
                    if fname not in incoming:
                        skipped.append(f"{fname}（未アップロード）")
                        continue
                    try:
                        df_imp = _read_csv_bytes(incoming[fname])
                        missing = [c for c in cols if c not in df_imp.columns]
                        if missing:
                            errors.append(f"{fname}: 必須列が不足 {missing}")
                            continue
                        _write_atomic_csv(df_imp[cols], path, cols)
                        applied.append(fname)
                    except Exception as e:
                        errors.append(f"{fname}: 取込エラー {e}")

                if applied: st.success("置換したファイル：" + " / ".join(applied))
                if skipped: st.info("スキップ：" + " / ".join(skipped))
                if errors:  st.error("エラー：" + " / ".join(errors))
                if applied:
                    time.sleep(1.2); st.rerun()

        # データ初期化
        with st.expander("🧯 データ初期化（ヘッダーのみ残す）", expanded=False):
            st.warning("⚠️ 取り消しできません。実行前に必ず『バックアップ』を取得してください。")
            tgt_att   = st.checkbox("勤怠データ（attendance_log.csv）を初期化", value=False)
            tgt_hreq  = st.checkbox("休日申請（holiday_requests.csv）を初期化", value=False)
            tgt_audit = st.checkbox("監査ログ（holiday_audit_log.csv）を初期化", value=False)
            tgt_login = st.checkbox("社員ログイン情報（社員ログイン情報.csv）も初期化（通常はOFF推奨）", value=False)

            confirm_text = st.text_input("確認のため 'DELETE' と入力してください", value="")
            do_init = st.button("🧨 初期化を実行", type="primary", disabled=(confirm_text.strip().upper() != "DELETE"))

            if do_init:
                try:
                    buf = io.BytesIO()
                    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
                        for path, cols, fname in BACKUP_TABLES:
                            dfb = _read_existing_or_empty(path, cols)
                            content = dfb[cols].to_csv(index=False)
                            zf.writestr(fname, content.encode("cp932"))
                    backup_dir = os.path.join(DATA_DIR, "backups")
                    os.makedirs(backup_dir, exist_ok=True)
                    backup_path = os.path.join(backup_dir, f"pre_wipe_{datetime.now():%Y%m%d_%H%M%S}.zip")
                    with open(backup_path, "wb") as f:
                        f.write(buf.getvalue())
                    st.info(f"既存データのバックアップを保存しました：{backup_path}")
                except Exception as e:
                    st.warning(f"バックアップで警告：{e}")

                done = []
                if tgt_att:
                    _write_atomic_csv(pd.DataFrame(columns=ATT_COLUMNS), CSV_PATH, ATT_COLUMNS); done.append("attendance_log.csv")
                if tgt_hreq:
                    _write_atomic_csv(pd.DataFrame(columns=HOLIDAY_COLUMNS), HOLIDAY_CSV, HOLIDAY_COLUMNS); done.append("holiday_requests.csv")
                if tgt_audit:
                    _write_atomic_csv(pd.DataFrame(columns=AUDIT_COLUMNS), AUDIT_LOG_CSV, AUDIT_COLUMNS); done.append("holiday_audit_log.csv")
                if tgt_login:
                    _write_atomic_csv(pd.DataFrame(columns=LOGIN_COLUMNS), LOGIN_CSV, LOGIN_COLUMNS); done.append("社員ログイン情報.csv")

                if done:
                    st.success("初期化完了：" + " / ".join(done))
                    time.sleep(1.2); st.rerun()
                else:
                    st.info("初期化対象が選択されていません。")

    # 社員UIへ進ませない
    st.stop()

# ==============================
# 社員UI
# ==============================
# ▼ サイドバーの切替メニュー（これでページを切替）
menu = st.sidebar.radio(
    "📑 表示メニュー",
    ["出退勤入力", "月別履歴", "休日申請"],
    index=0,
    key="main_view_selector"
)

if menu == "出退勤入力":
    st.header("📝 出退勤の入力")

    # === 入力可能な過去期間の設定（例：直近2ヶ月） ===
    PAST_MONTHS = 2
    today = today_jst()
    try:
        from dateutil.relativedelta import relativedelta
        past_limit_date = today - relativedelta(months=PAST_MONTHS)
    except Exception:
        past_limit_date = today - timedelta(days=31*PAST_MONTHS)

    # ここからタブ
    tab_punch, tab_edit, tab_ot = st.tabs(["📝 打刻入力", "✏️ 修正／削除", "⏱️ 残業申請"])

    with tab_punch:
        # 社員UI：日付入力（前月ロックのUX強化）
        punch_type = st.radio("打刻種類を選択", ["出勤", "退勤"], horizontal=True)
        selected_date = st.date_input(
            "日付",
            value=today,
            min_value=past_limit_date,     # ← 直近◯ヶ月まで遡れる
            max_value=today                # ← 未来は不可
        )

        # ---- 打刻抑止：承認済み休日なら保存ボタンを無効化 ----
        holiday_df_all = read_holiday_csv()
        sel_date_str = selected_date.strftime("%Y-%m-%d")
        is_approved_holiday = bool((
            (holiday_df_all["社員ID"] == st.session_state.user_id) &
            (holiday_df_all["休暇日"] == sel_date_str) &
            (holiday_df_all["ステータス"] == "承認")
        ).any())

        # ========= 背景GPS取得（UI＋非表示JS）=========

        # セッション初期化
        if "manual_gps" not in st.session_state:
            st.session_state.manual_gps = ""   # "lat,lng"
        if "gps_error" not in st.session_state:
            st.session_state.gps_error = ""
        if "gps_click_token" not in st.session_state:
            st.session_state.gps_click_token = 0.0  # ボタン押下トリガ

        # ===== ここから “ギャップを詰めたい範囲” を本物の親で囲む =====
        with st.container():
            col_g1, col_g2 = st.columns([1, 3])
            with col_g1:
                # 押下でトークン更新→即 rerun（JS が新トークンを拾ってポップアップ起動）
                if st.button("位置情報を取得する"):
                    st.session_state.gps_error = ""
                    st.session_state.manual_gps = ""
                    st.session_state.gps_click_token = time.time()
                    st.rerun()

                # 保存ボタン
                save_clicked = st.button("保存", key="save_btn_top", disabled=is_approved_holiday)

                # ▼ 未申請の残業アラート（当月：start_date～end_date） ← 保存ボタンの直下に出す
                try:
                    att_period = df[
                        (df["社員ID"] == st.session_state.user_id) &
                        (df["日付"] >= start_date) & (df["日付"] <= end_date)
                    ].copy()
                    att_period["残業時間"] = att_period["残業時間"].astype(float)

                    # 自動計算で残業>0の日
                    overtime_dates = set(
                        att_period.loc[att_period["残業時間"] > 0, "日付"].dt.strftime("%Y-%m-%d")
                    )

                    # すでに「申請済 or 承認」の対象日
                    ot_all = read_overtime_csv()
                    applied_dates = set(
                        ot_all[
                            (ot_all["社員ID"] == st.session_state.user_id) &
                            (ot_all["対象日"] >= start_date.strftime("%Y-%m-%d")) &
                            (ot_all["対象日"] <= end_date.strftime("%Y-%m-%d")) &
                            (ot_all["ステータス"].isin(["申請済", "承認"]))
                        ]["対象日"].tolist()
                    )

                    pending_unapplied = sorted(overtime_dates - applied_dates)
                    if pending_unapplied:
                        ex = "、".join(pending_unapplied[:3]) + (" など" if len(pending_unapplied) > 3 else "")
                        st.info(f"⚠️ 未申請の残業があります。『⏱️ 残業申請』タブから申請してください。例：{ex}")
                except Exception:
                    pass

            with col_g2:
                # 現状表示
                if st.session_state.manual_gps:
                    st.success(f"取得済み: {st.session_state.manual_gps}")
                elif st.session_state.gps_error:
                    st.warning("取得失敗: " + st.session_state.gps_error)
                else:
                    st.caption("未取得です（位置情報を取得してください）")

            # ---- geolocation 実行用（keyは渡さない）----
            TOKEN_VAL = float(st.session_state.get("gps_click_token", 0) or 0)

            # ★ トークンが有効な時だけコンポーネントを描画する
            if TOKEN_VAL > 0:
                st.markdown('<div class="g-cmark"></div>', unsafe_allow_html=True)
                gps_val = components.html(
                    """
                    <div id="gps-hook" style="display:none"></div>
                    <script>
                    (function(){
                      const TOKEN = "__TOKEN__";
                      if (!TOKEN || TOKEN === "0" || TOKEN === "0.0") return;

                      function redirectWith(param, value){
                        try {
                          const topWin = window.top;
                          const url = new URL(topWin.location.href);
                          url.searchParams.set(param, value);
                          topWin.location.href = url.toString();
                        } catch (e) {}
                      }
                      let w = window.open("", "_blank", "width=360,height=280");
                      if (!w) { redirectWith("gps_error","POPUP_BLOCKED"); return; }

                      w.document.write(`<!doctype html><html><head>
                        <meta name="viewport" content="width=device-width,initial-scale=1"/>
                        <title>位置情報の取得</title>
                      </head>
                      <body style="font-family:system-ui,-apple-system,Segoe UI,Roboto; padding:1rem">
                        <div style="margin-bottom:0.75rem;">位置情報を取得しています…<br>ブラウザの許可ダイアログを確認してください。</div>
                        <div id="s" style="white-space:pre-wrap"></div>
                        <script>
                          (function(){
                            const say = (t) => { try { document.getElementById('s').textContent = t; } catch (_) {} };
                            function back(param, value){
                              try{
                                const topWin = window.opener ? window.opener.top : null;
                                if (topWin){
                                  const url = new URL(topWin.location.href);
                                  url.searchParams.set(param, value);
                                  topWin.location.href = url.toString();
                                }
                              }catch(e){}
                              setTimeout(()=>window.close(), 300);
                            }

                            if (!('geolocation' in navigator)) { say("この端末/ブラウザでは位置情報が使えません。"); back("gps_error","GEO_UNSUPPORTED"); return; }

                            navigator.geolocation.getCurrentPosition(function(pos){
                              const v = pos.coords.latitude + "," + pos.coords.longitude;
                              say("取得成功: " + v + "（このウィンドウは自動で閉じます）");
                              back("gps", v);
                            }, function(err){
                              const msg = (err && err.message) ? err.message : "GEO_ERROR";
                              say("取得失敗: " + msg + "（このウィンドウは自動で閉じます）");
                              back("gps_error", msg);
                            }, { enableHighAccuracy:true, timeout:15000, maximumAge:0 });
                          })();
                        <\/script>
                      </body></html>`);
                    })();
                    </script>
                    """.replace("__TOKEN__", str(TOKEN_VAL)),
                    height=0
                )

                # 値の受け取り（同じ）
                if isinstance(gps_val, str) and gps_val:
                    if gps_val.startswith("ERROR:"):
                        st.session_state.gps_error = gps_val.replace("ERROR:", "")
                        st.session_state.manual_gps = ""
                    else:
                        st.session_state.manual_gps = gps_val
                        st.session_state.gps_error = ""
                    st.session_state.gps_click_token = 0
                    st.rerun()
            # else: TOKENが0の時は何も描画しない（＝空白が一切できない）


            # Python側で使う値（保存処理で使用）
            effective_gps = st.session_state.get("manual_gps", "")
            lat, lng = "", ""
            if isinstance(effective_gps, str) and "," in effective_gps:
                lat, lng = [s.strip() for s in effective_gps.split(",", 1)]

            # ========= 背景GPS取得ここまで =========

            # …（gps_val の処理、lat/lng の算出の直後あたりに）
            if save_clicked:
                if punch_type == "出勤" and not (lat and lng):
                    err = st.session_state.get("gps_error", "")
                    st.warning("位置情報が未取得のため、位置情報なしで保存します。"
                               + (f"（原因: {err.replace('ERROR:','')}）" if err else ""))

                st.session_state.pending_save = True
                st.session_state.punch_action = {
                    "type": punch_type,
                    "date": selected_date.strftime("%Y-%m-%d"),
                }
                st.rerun()

            # ---- 前月ロック判定 ----
            if selected_date < past_limit_date or selected_date > today:
                st.error(f"この日は入力範囲外です。{past_limit_date:%Y-%m-%d} 〜 {today:%Y-%m-%d} の間で選択してください。")
            else:
                # （任意）承認済み休日なら事前に注意を表示
                if is_approved_holiday:
                    st.warning("この日は承認済みです。打刻する場合は、管理者にご相談ください。")

                # ===== 保存フェーズ（pending_save が True のときに実行） =====
                if st.session_state.get("pending_save"):
                    # 今ランでは位置情報取得は起動しない（任意のため）
                    action = st.session_state.get("punch_action", {})
                    action_type = action.get("type", punch_type)  # 念のためフォールバック
                    action_date = action.get("date", selected_date.strftime("%Y-%m-%d"))
                    now_hm = datetime.now(JST).strftime("%H:%M")

                    # 承認済み休日は保存禁止（仕様）
                    _hd = read_holiday_csv()
                    if ((_hd["社員ID"] == st.session_state.user_id) &
                        (_hd["休暇日"] == action_date) &
                        (_hd["ステータス"] == "承認")).any():
                        st.session_state.pending_save = False
                        st.error("この日は承認済みの休日です。打刻はできません。")
                        st.stop()

                    # 保存本体（出勤/退勤 共通）
                    df_att = _read_csv_flexible(CSV_PATH) if os.path.exists(CSV_PATH) else pd.DataFrame(columns=ATT_COLUMNS)
                    for col in ATT_COLUMNS:
                        if col not in df_att.columns:
                            df_att[col] = ""

                    m = (df_att["社員ID"] == st.session_state.user_id) & (df_att["日付"] == action_date)

                    if action_type == "出勤":
                        # 位置情報が無い場合は警告だけ出して保存続行
                        if not (lat and lng):
                            err = st.session_state.get("gps_error", "")
                            st.warning("位置情報が未取得のため、位置情報なしで保存します。"
                                       + (f"（原因: {err.replace('ERROR:','')}）" if err else ""))
                        if m.any():
                            df_att.loc[m, ["出勤時刻", "緯度", "経度"]] = [now_hm, (lat or ""), (lng or "")]
                        else:
                            df_att = pd.concat([df_att, pd.DataFrame([{
                                "社員ID": st.session_state.user_id, "氏名": st.session_state.user_name,
                                "日付": action_date, "出勤時刻": now_hm, "退勤時刻": "",
                                "緯度": (lat or ""), "経度": (lng or "")
                            }])], ignore_index=True)

                        if safe_write_csv(df_att, CSV_PATH, ATT_COLUMNS):
                            removed = auto_cancel_holiday_by_attendance(st.session_state.user_id, st.session_state.user_name, action_date)
                            if removed > 0:
                                st.info(f"この日の休暇申請（{removed}件）を自動取消しました。")
                            st.session_state.pending_save = False
                            st.success(f"✅ 出勤 を {now_hm} で保存しました。")
                            time.sleep(1.2)
                            st.rerun()

                    else:  # 退勤
                        if m.any():
                            # 座標があれば一緒に更新、無ければ退勤時刻のみ
                            if lat and lng:
                                df_att.loc[m, ["退勤時刻", "緯度", "経度"]] = [now_hm, lat, lng]
                            else:
                                df_att.loc[m, "退勤時刻"] = now_hm
                        else:
                            # 新規行（退勤先行）。座標があれば入れる
                            df_att = pd.concat([df_att, pd.DataFrame([{
                                "社員ID": st.session_state.user_id, "氏名": st.session_state.user_name,
                                "日付": action_date, "出勤時刻": "", "退勤時刻": now_hm,
                                "緯度": (lat if (lat and lng) else ""), "経度": (lng if (lat and lng) else "")
                            }])], ignore_index=True)

                        if safe_write_csv(df_att, CSV_PATH, ATT_COLUMNS):
                            st.session_state.pending_save = False
                            st.success(f"✅ 退勤 を {now_hm} で保存しました。")
                            time.sleep(1.2)
                            st.rerun()

    # ==============================
    # 修正 / 削除（社員本人のみ）
    # ==============================
    with tab_edit:
        with st.expander("出退勤の ✏️ 修正 / 🗑️ 削除", expanded=False):
            df_self = df[
                (df["社員ID"] == st.session_state.user_id) &
                (df["日付"] >= start_date) & (df["日付"] <= end_date) &
                (df["日付"] >= OPEN_START)  # 当月以降のみ編集可
            ].sort_values("日付")

            if df_self.empty:
                st.caption("当月データがありません。")
            else:
                choice_dates = df_self["日付"].dt.strftime("%Y-%m-%d").tolist()
                colL, colR = st.columns(2)
                with colL:
                    edit_date_str = st.selectbox("修正する日付を選択", options=choice_dates, key="self_edit_date")
                row_cur = df_self[df_self["日付"].dt.strftime("%Y-%m-%d") == edit_date_str].iloc[0]
                with colR:
                    st.caption(f"選択中：{row_cur['氏名']} / {edit_date_str}")

                c1, c2, c3 = st.columns([1,1,1])
                with c1:
                    new_start = st.text_input("出勤（HH:MM）", value=str(row_cur["出勤時刻"] or ""), key="self_edit_start")
                with c2:
                    new_end   = st.text_input("退勤（HH:MM）", value=str(row_cur["退勤時刻"] or ""), key="self_edit_end")
                with c3:
                    if st.button("この日の時刻を更新", key="self_edit_apply"):
                        def _ok(t):
                            if not str(t).strip(): return True
                            try:
                                datetime.strptime(str(t).strip(), "%H:%M")
                                return True
                            except:
                                return False
                        if not (_ok(new_start) and _ok(new_end)):
                            st.error("時刻は HH:MM 形式で入力してください（例：07:30）。")
                        else:
                            df_all = _read_csv_flexible(CSV_PATH).fillna("")
                            m = (df_all["社員ID"]==st.session_state.user_id) & (df_all["日付"]==edit_date_str)
                            if not m.any():
                                st.warning("該当日の記録が見つかりませんでした。")
                            else:
                                df_all.loc[m, "出勤時刻"] = str(new_start).strip()
                                df_all.loc[m, "退勤時刻"] = str(new_end).strip()
                                if safe_write_csv(df_all, CSV_PATH, ATT_COLUMNS):
                                    dept_me = (df_login.loc[df_login["社員ID"]==st.session_state.user_id, "部署"].iloc[0]
                                               if (df_login["社員ID"]==st.session_state.user_id).any() else "")
                                    try:
                                        _base = pd.Timestamp.today().normalize()
                                        start_dt = pd.to_datetime(new_start, format="%H:%M", errors="coerce")
                                        end_dt   = pd.to_datetime(new_end,   format="%H:%M", errors="coerce")
                                        rec = {
                                            "出_dt": pd.Timestamp.combine(_base.date(), start_dt.time()) if pd.notna(start_dt) else pd.NaT,
                                            "退_dt": pd.Timestamp.combine(_base.date(), end_dt.time())   if pd.notna(end_dt)   else pd.NaT,
                                            "部署": dept_me
                                        }
                                        work_h, ot_h = calc_work_overtime(rec)
                                        st.success(f"更新しました。参考：勤務 {format_hours_minutes(work_h)} / 残業 {format_hours_minutes(ot_h)}")
                                    except:
                                        st.success("更新しました。残業は一覧再描画時に自動再計算されます。")
                                    time.sleep(1)
                                st.rerun()

                st.markdown("—")
                st.markdown("#### 🗑️ 削除（複数選択可）")
                del_df = df_self.copy()
                del_df["日付"] = del_df["日付"].dt.strftime("%Y-%m-%d")
                del_df = del_df[["日付","出勤時刻","退勤時刻"]].assign(削除=False)

                edited = st.data_editor(
                    del_df,
                    use_container_width=True,
                    num_rows="fixed",
                    hide_index=True,
                    key="self_delete_editor",
                    column_config={
                        "削除": st.column_config.CheckboxColumn("削除", help="削除する行にチェック"),
                        "日付": st.column_config.TextColumn("日付", disabled=True),
                        "出勤時刻": st.column_config.TextColumn("出勤時刻", disabled=True),
                        "退勤時刻": st.column_config.TextColumn("退勤時刻", disabled=True),
                    }
                )

                to_delete = edited.loc[edited["削除"]==True, "日付"].tolist()
                colA, colB = st.columns([1,2])
                with colA:
                    confirm_del = st.checkbox("本当に削除しますか？", key="self_delete_confirm")
                with colB:
                    if st.button("選択した行を削除", disabled=(len(to_delete)==0 or not confirm_del), key="self_delete_apply"):
                        df_all = _read_csv_flexible(CSV_PATH).fillna("")
                        for d in to_delete:
                            mask = (df_all["社員ID"]==st.session_state.user_id) & (df_all["日付"]==d)
                            df_all = df_all[~mask]
                        if safe_write_csv(df_all, CSV_PATH, ATT_COLUMNS):
                            st.success(f"{len(to_delete)} 件削除しました。")
                            time.sleep(1)
                            st.rerun()
    # ==============================
    # 残業申請
    # ==============================
    with tab_ot:
        with st.expander("申請フォーム", expanded=True):
            with st.form("overtime_form"):
                target_date = st.date_input(
                    "対象日",
                    value=today,
                    min_value=OPEN_START.date(),
                    max_value=OPEN_END.date(),
                    key="ot_target_date"
                )

                c1, c2 = st.columns(2)
                with c1:
                    start_str = st.text_input("残業 開始（HH:MM）", value="", key="ot_start_hhmm")
                with c2:
                    end_str   = st.text_input("残業 終了（HH:MM）", value="", key="ot_end_hhmm")

                reason = st.text_input("申請理由（任意だが推奨）", value="", key="ot_reason")
                submitted = st.form_submit_button("申請する", type="primary")

                # 入力のプレビュー（形式が正しければ所要時間を表示）
                if _is_hhmm(start_str) and _is_hhmm(end_str):
                    _base = pd.Timestamp.today().normalize()
                    s = pd.to_datetime(start_str, format="%H:%M")
                    e = pd.to_datetime(end_str,   format="%H:%M")
                    sdt = pd.Timestamp.combine(_base.date(), s.time())
                    edt = pd.Timestamp.combine(_base.date(), e.time())
                    if edt > sdt:
                        mins = int((edt - sdt).total_seconds() // 60)
                        hrs_f = round(mins / 60.0, 2)
                        st.caption(f"⏱️ 申請時間：{mins}分（= {hrs_f} 時間）")
                    else:
                        st.caption("終了は開始より後にしてください。")

                if submitted:
                    # バリデーション
                    if not (_is_hhmm(start_str) and _is_hhmm(end_str)):
                        st.error("開始・終了は HH:MM 形式で入力してください（例：18:00）。")
                        st.stop()

                    _base = pd.Timestamp.today().normalize()
                    s = pd.to_datetime(start_str, format="%H:%M", errors="coerce")
                    e = pd.to_datetime(end_str,   format="%H:%M", errors="coerce")
                    if pd.isna(s) or pd.isna(e):
                        st.error("開始・終了の時刻が不正です。")
                        st.stop()
                    sdt = pd.Timestamp.combine(_base.date(), s.time())
                    edt = pd.Timestamp.combine(_base.date(), e.time())
                    if not (edt > sdt):
                        st.error("終了は開始より後にしてください。")
                        st.stop()

                    mins = int((edt - sdt).total_seconds() // 60)
                    if mins <= 0:
                        st.error("申請時間は1分以上にしてください。")
                        st.stop()

                    hrs_f = round(mins / 60.0, 2)  # CSVには従来通り「時間(小数)」で保存
                    _dstr = target_date.strftime("%Y-%m-%d")

                    ot = read_overtime_csv()
                    dup_mask = (
                        (ot["社員ID"] == st.session_state.user_id) &
                        (ot["対象日"] == _dstr) &
                        (ot["ステータス"].isin(["申請済", "承認"]))
                    )
                    if dup_mask.any():
                        st.warning("この日付は、すでに『申請中』または『承認済』の残業申請があります。")
                    else:
                        new_row = {
                            "社員ID": st.session_state.user_id,
                            "氏名": st.session_state.user_name,
                            "対象日": _dstr,
                            "申請日時": datetime.now(JST).strftime("%Y-%m-%d %H:%M:%S"),
                            "申請残業H": f"{hrs_f:.2f}",   # ← 小数時間で保存
                            "申請理由": reason,
                            "ステータス": "申請済",
                            "承認者": "", "承認日時": "", "却下理由": ""
                        }
                        ot = pd.concat([ot, pd.DataFrame([new_row])], ignore_index=True)
                        write_overtime_csv(ot)
                        st.success(f"✅ 残業申請を受け付けました（{mins}分）。")
                        time.sleep(1); st.rerun()

# ==============================
# 月別履歴（社員）
# ==============================
if menu == "月別履歴":
    st.header(f"📋 月別履歴（{start_date:%Y/%m/%d}～{end_date:%Y/%m/%d}）")

    df_self = df[
        (df["社員ID"] == st.session_state.user_id) &
        (df["日付"] >= start_date) &
        (df["日付"] <= end_date)
    ].sort_values("日付")

    if df_self.empty:
        st.info("この月の出退勤記録はありません。")
    else:
        df_view = df_self.copy()
        df_view["日付"] = df_view["日付"].dt.strftime("%Y-%m-%d")
        df_view = df_view.rename(columns={"日付":"日付","出勤時刻":"出勤","退勤時刻":"退勤","残業時間":"残業H"})
        if "残業H" in df_view.columns:
            df_view["残業H"] = df_view["残業H"].astype(float).apply(format_hours_minutes)
        if "承認残業時間" in df_view.columns:
            df_view["残業H(承認)"] = df_view["承認残業時間"].astype(float).apply(format_hours_minutes)

        cols = ["日付", "出勤", "退勤"]
        if "残業H" in df_view.columns:
            cols.append("残業H")
        if "承認残業時間" in df_view.columns:
            cols.append("残業H(承認)")

        # ▼ 1ページの件数
        per_page = st.selectbox("1ページの件数", [10, 20, 30, 50, 100], index=0, key="mh_per_page")
        # ▼ ページネーションして表示
        paged, _, _ = paginate_df(df_view[cols].reset_index(drop=True), page_key="mh_page", per_page=int(per_page))
        st.dataframe(paged, hide_index=True, use_container_width=True)

        # 合計の表示
        st.markdown(f"**🕒 合計残業時間（総計）：{format_hours_minutes(df_self['残業時間'].sum())}**")
        if "承認残業時間" in df_self.columns:
            st.markdown(f"**✅ 合計残業時間（承認済み）：{format_hours_minutes(df_self['承認残業時間'].sum())}**")

# ==============================
# 休日・休暇申請
# ==============================
if menu == "休日申請":
    st.header("📅 休日・休暇申請")

    # 申請フォーム（そのまま流用）
    with st.form("holiday_form"):
        holiday_date = st.date_input("休暇日", value= today_jst(), min_value=today_jst())
        holiday_type = st.selectbox("休暇種類", ["希望休", "特別休暇（冠婚葬祭など）", "その他（備考有り）"])
        notes = st.text_input("備考（その他の理由）") if holiday_type == "その他（備考有り）" else ""
        submitted = st.form_submit_button("申請する", type="primary")
        if submitted:
            df_holiday = read_holiday_csv()
            new_record = {
                "社員ID":  st.session_state.user_id,
                "氏名":    st.session_state.user_name,
                "申請日":  today_jst().strftime("%Y-%m-%d"),
                "休暇日":  holiday_date.strftime("%Y-%m-%d"),
                "休暇種類": holiday_type,
                "備考":    notes,
                "ステータス": "申請済",
                "承認者": "", "承認日時": "", "却下理由": ""
            }
            df_holiday = pd.concat([df_holiday, pd.DataFrame([new_record])], ignore_index=True)
            write_holiday_csv(df_holiday)
            st.success("✅ 休暇申請を受け付けました")
            time.sleep(1); st.rerun()

    # 当月の申請一覧（ページネーション付き）
    hd = read_holiday_csv()
    month_mask = (
        (hd["社員ID"] == st.session_state.user_id) &
        (hd["休暇日"] >= start_date.strftime("%Y-%m-%d")) &
        (hd["休暇日"] <= end_date.strftime("%Y-%m-%d"))
    )
    hd_month = hd.loc[month_mask, ["休暇日", "休暇種類", "ステータス", "承認者", "承認日時", "却下理由"]] \
                .sort_values("休暇日")

    st.subheader("当月の申請一覧")
    if hd_month.empty:
        st.caption("この期間の申請はありません。")
    else:
        show = hd_month.rename(columns={"休暇日":"日付","休暇種類":"区分","ステータス":"状態"}).reset_index(drop=True)
        per_page_h = st.selectbox("1ページの件数（申請一覧）", [10, 20, 30, 50, 100], index=0, key="hol_per_page")
        paged_h, _, _ = paginate_df(show, page_key="hol_page", per_page=int(per_page_h))
        st.dataframe(paged_h, hide_index=True, use_container_width=True)

    # 申請済の取消（本人）—（元のまま＋必要ならページネーション）
    st.subheader("申請済の取消（本人）")
    hd_all_my = read_holiday_csv()
    cand = hd_all_my[
        (hd_all_my["社員ID"] == st.session_state.user_id) &
        (hd_all_my["ステータス"] == "申請済")
    ].copy() if not hd_all_my.empty else pd.DataFrame(columns=HOLIDAY_COLUMNS)

    if cand.empty:
        st.caption("取消できる申請はありません（申請済が無いか、すでに承認/却下済みです）。")
    else:
        cand = cand.sort_values(["休暇日","申請日"])
        view_cancel = cand[["休暇日","休暇種類","申請日","備考"]].copy().rename(
            columns={"休暇日":"日付","休暇種類":"区分"}
        ).reset_index(drop=True)

        # ページネーション（取消候補）
        per_page_c = st.selectbox("1ページの件数（取消候補）", [10, 20, 30, 50, 100], index=0, key="hol_cancel_per_page")
        paged_c, _, _ = paginate_df(view_cancel, page_key="hol_cancel_page", per_page=int(per_page_c))

        # data_editor はチェック列を足す必要があるので、表示対象のページ分だけを編集
        paged_c = paged_c.copy()
        paged_c["取消"] = False
        edited_cancel = st.data_editor(
            paged_c,
            hide_index=True, use_container_width=True,
            column_config={
                "日付": st.column_config.TextColumn("日付", disabled=True),
                "区分": st.column_config.TextColumn("休暇種類", disabled=True),
                "申請日": st.column_config.TextColumn("申請日", disabled=True),
                "備考": st.column_config.TextColumn("備考", disabled=True),
                "取消": st.column_config.CheckboxColumn("この申請を取り消す"),
            },
            key="self_cancel_pending_holidays_paged"
        )

        # 取消対象（表示中ページ）を元DFのキーに戻す
        to_cancel = []
        for _, r in edited_cancel[edited_cancel["取消"]==True].iterrows():
            to_cancel.append([r["日付"], r["申請日"]])

        if st.button("選択した『申請済』を取消", key="hol_cancel_button"):
            if not to_cancel:
                st.info("取り消す行が選択されていません。")
            else:
                base = read_holiday_csv()
                before = len(base)
                rows_for_audit = []
                when_ts = datetime.now(JST).strftime("%Y-%m-%d %H:%M:%S")
                for d, applied_on in to_cancel:
                    km = (
                        (base["社員ID"] == st.session_state.user_id) &
                        (base["休暇日"] == d) &
                        (base["申請日"] == applied_on) &
                        (base["ステータス"] == "申請済")
                    )
                    if km.any():
                        rows_for_audit.append({
                            "timestamp": when_ts, "承認者": st.session_state.user_name,
                            "社員ID": st.session_state.user_id, "氏名": st.session_state.user_name,
                            "休暇日": d, "申請日": applied_on,
                            "旧ステータス": "申請済", "新ステータス": "本人取消", "却下理由": ""
                        })
                        base = base[~km]
                write_holiday_csv(base)
                append_audit_log(rows_for_audit)
                st.success(f"{before-len(base)} 件の『申請済』を取り消しました。")
                time.sleep(1); st.rerun()



st.caption("build: 2025-09-01 22:15 JPY v2")
